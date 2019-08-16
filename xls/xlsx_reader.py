# author:	Ivan Kovalev
# read xlsx file and obtain data to generate configs
# Excel fields:
# 0)	№
# 1)	Район
# 2)	Модель
# 3)	Тип УИВ
# 4)	hostname
# 5)	eth0 - порт1
# 6)	eth1 - порт2
# 	bond0.100
# 7)		Первый байт
# 8)		Второй байт
# 9)		Третий байт
# 10)		Четвертый байт
# 11)		Маска
# 	bond0.101
# 12)		Первый байт
# 13)		Второй байт
# 14)		Третий байт
# 15)		Четвертый байт
# 16)		Маска
# 	bond0.102
# 17)		Первый байт
# 18)		Второй байт
# 19)		Третий байт
# 20)		Четвертый байт
# 21)		Маска
# 	eth2 - порт4 - ISP 1 коммутатор 1
# 22)		Первый байт
# 23)		Второй байт
# 24)		Третий байт
# 25)		Четвертый байт
# 26)		Маска
# 	eth3 - порт3 - ISP 2 коммутатор 2
# 27)		Первый байт
# 28)		Второй байт
# 29)		Третий байт
# 30)		Четвертый байт
# 31)		Маска
# 	eth4 - SFP
# 32)		Первый байт
# 33)		Второй байт
# 34)		Третий байт
# 35)		Четвертый байт
# 36)		Маска
# 	failover bond0.100
# 37)		Первый байт
# 38)		Второй байт
# 39)		Третий байт
# 40)		Четвертый байт
# 	failover bond0.101
# 41)		Первый байт
# 42)		Второй байт
# 43)		Третий байт
# 44)		Четвертый байт
# 	failover bond0.102
# 45)		Первый байт
# 46)		Второй байт
# 47)		Третий байт
# 48)		Четвертый байт
# 	failover ISP1
# 49)		Первый байт
# 50)		Второй байт
# 51)		Третий байт
# 52)		Четвертый байт
# 	failover ISP2
# 53)		Первый байт
# 54)		Второй байт
# 55)		Третий байт
# 56)		Четвертый байт
# 	Туннелируемые сети
# 57)		Первый байт
# 58)		Второй байт
# 59)		Третий байт
# 60)		Четвертый байт
# 61	Пароль enable
import openpyxl
import glog as log
from globals import general_params


def __parse_row(row):
    return {
        'number': row[0].value,
        'region': row[1].value,
        'model': row[2].value,
        'type': row[3].value,
        'hostname': row[4].value,
        'eth0-port1': row[5].value,
        'eth1-port2': row[6].value,
        'bond0.100-b1': row[7].value,
        'bond0.100-b2': row[8].value,
        'bond0.100-b3': row[9].value,
        'bond0.100-b4': row[10].value,
        'bond0.100-mask': row[11].value,
        'bond0.101-b1': row[12].value,
        'bond0.101-b2': row[13].value,
        'bond0.101-b3': row[14].value,
        'bond0.101-b4': row[15].value,
        'bond0.101-mask': row[16].value,
        'bond0.102-b1': row[17].value,
        'bond0.102-b2': row[18].value,
        'bond0.102-b3': row[19].value,
        'bond0.102-b4': row[20].value,
        'bond0.102-mask': row[21].value,
        'eth2-p4-isp1-c1-b1': row[22].value,
        'eth2-p4-isp1-c1-b2': row[23].value,
        'eth2-p4-isp1-c1-b3': row[24].value,
        'eth2-p4-isp1-c1-b4': row[25].value,
        'eth2-p4-isp1-c1-mask': row[26].value,
        'eth3-p3-isp2-c2-b1': row[27].value,
        'eth3-p3-isp2-c2-b2': row[28].value,
        'eth3-p3-isp2-c2-b3': row[29].value,
        'eth3-p3-isp2-c2-b4': row[30].value,
        'eth3-p3-isp2-c2-mask': row[31].value,
        'eth4-sfp-b1': row[32].value,
        'eth4-sfp-b2': row[33].value,
        'eth4-sfp-b3': row[34].value,
        'eth4-sfp-b4': row[35].value,
        'eth4-sfp-mask': row[36].value,
        'failover-bond0.100-b1': row[37].value,
        'failover-bond0.100-b2': row[38].value,
        'failover-bond0.100-b3': row[39].value,
        'failover-bond0.100-b4': row[40].value,
        'failover-bond0.101-b1': row[41].value,
        'failover-bond0.101-b2': row[42].value,
        'failover-bond0.101-b3': row[43].value,
        'failover-bond0.101-b4': row[44].value,
        'failover-bond0.102-b1': row[45].value,
        'failover-bond0.102-b2': row[46].value,
        'failover-bond0.102-b3': row[47].value,
        'failover-bond0.102-b4': row[48].value,
        'failover-isp1-b1': row[49].value,
        'failover-isp1-b2': row[50].value,
        'failover-isp1-b3': row[51].value,
        'failover-isp1-b4': row[52].value,
        'failover-isp2-b1': row[53].value,
        'failover-isp2-b2': row[54].value,
        'failover-isp2-b3': row[55].value,
        'failover-isp2-b4': row[56].value,
        'tunnel-nets-b1': row[57].value,
        'tunnel-nets-b2': row[58].value,
        'tunnel-nets-b3': row[59].value,
        'tunnel-nets-b4': row[60].value,
        'pass_enable': row[61].value,
    }


def get_config_data():
    log.info("xls::get_config_data")
    res = []
    wb = openpyxl.load_workbook(general_params.input_file_path)
    ws = wb.get_sheet_by_name(wb.sheetnames[0])
    for i, row in enumerate(ws.iter_rows()):
        if i < 2:
            # skip table header
            continue
        data = __parse_row(row)
        can_add = True
        for k, v in data.items():
            if k != 'pass_enable' and not v:
                log.info(f"xls::get_config_data :: row {i} skipped because of empty value in one of the columns")
                can_add = False
                break
        can_add and res.append(data)

    log.info(f"xls::get_config_data :: done {len(res)} rows has been read")
    return res
